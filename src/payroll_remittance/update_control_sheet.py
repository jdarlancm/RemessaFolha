from openpyxl import load_workbook
from pypdf import PdfReader

from helpers import paths_helper, paycheck_helper
from utils import os_utils, date_utils


def update(month, year):
    payroll_salaries = _extract_payroll_salaries(month, year)
    _update_payroll_sheet(month, year, payroll_salaries)
    print(payroll_salaries)


def _extract_payroll_salaries(month, year):
    paychecks_path = paths_helper.get_payroll_receipts_folder(month, year)
    payroll_salaries = []

    paycheck_list = os_utils.list_files(paychecks_path, ".pdf")

    for paycheck_file in paycheck_list:
        paycheck_complete_filename = f"{paychecks_path}\\{paycheck_file}"
        matricula, salario_liquido = _extract_employee_salary(
            paycheck_complete_filename
        )
        payroll_salaries.append(
            {"matricula": matricula, "salario_liquido": salario_liquido}
        )

    return payroll_salaries


def _extract_employee_salary(paycheck_complete_filename):
    first_page = 0
    reader = PdfReader(paycheck_complete_filename)
    page_content = reader.pages[first_page].extract_text()
    # TODO checar se é um contra-cheque
    matricula = paycheck_helper.extract_matricula(page_content)
    salario_liquido = paycheck_helper.extract_net_salary(page_content)

    if not salario_liquido or matricula == 0:
        raise ValueError(
            f"Não foi possível obter informaçoes do funcionário no contra-cheque ({paycheck_complete_filename})"
        )

    return matricula, salario_liquido


def _update_payroll_sheet(month, year, payroll_salaries):
    COL_MATRICULA = 1

    payroll_filename = paths_helper.get_payroll_year_complete_filename(year)
    spreadsheet_tab = "folha"

    wb = load_workbook(filename=payroll_filename)
    spreadsheet = wb[spreadsheet_tab]

    col_reference_month = _get_reference_month(month, spreadsheet)

    if not col_reference_month:
        raise ValueError(
            f"Não foi encontrada a coluna de referência do mes {year}-{month} na planilha da folha: {payroll_filename}"
        )

    for row in spreadsheet.iter_rows(min_row=3, max_row=spreadsheet.max_row):
        matricula = row[COL_MATRICULA].value
        salario_liquido = _get_salary_from_payroll(matricula, payroll_salaries)
        if salario_liquido:
            cell_coordinate = f"{col_reference_month}{row[0].row}"
            spreadsheet[cell_coordinate] = float(salario_liquido)

    wb.save(payroll_filename)


def _get_reference_month(month, spreadsheet):
    ROW_HEADER = 2
    col_name_reference_month = f"{date_utils.nome_mes(month)}-fol"

    for cell in spreadsheet[ROW_HEADER]:
        if cell.value == col_name_reference_month:
            return cell.column_letter

    return None


def _get_salary_from_payroll(matricula, payroll_data):
    for employee in payroll_data:
        if int(employee["matricula"]) == matricula:
            return employee["salario_liquido"].replace(",", ".")

    return None
